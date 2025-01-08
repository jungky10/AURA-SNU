from matplotlib import pyplot as plt
import matplotlib
import pytz
import datetime as dt
import os
import matplotlib.dates as mdates
from matplotlib import gridspec
import pandas as pd
import openpyxl as xl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl import load_workbook
import sys
import io
from matplotlib import pyplot as plt
import datetime as dt
import openpyxl as xl
import math 
import numpy as np
import sys
import gc
import re
sys.setrecursionlimit(6000) 
matplotlib.use('Qt5Agg')

plt.style.use(['default'])

RWA_Event = ["Any","Phasic", "Intermediate", "Tonic"]
percentile = [50, 80, 95]
ev_index = {
    0 : [0,3,4], # Tonic
    1 : [2,4], # Intermediate
    2 : [1,3], # Phasic
    3 : [0,1,2,3,4] # Any
}

ev_index_AASM = {
    0 : [0], # Tonic
    1 : [1], # Phasic
    2 : [0,1,2], # Any
}

para = [
    "CRWA_T","DURmean_T","DURp50_T","DURp80_T","DURp95_T","RWAI_T",
    "CRWA_I","DURmean_I","DURp50_I","DURp80_I","DURp95_I","RWAI_I",
    "CRWA_P","DURmean_P","DURp50_P","DURp80_P","DURp95_P","RWAI_P",
    "CRWA_A","DURmean_A","DURp50_A","DURp80_A","DURp95_A","RWAI_A",
] # CRWA: cumulative RWA%, DUR x: x percentile duration of individual RWA in total (sec), RWAI: num of RWA per hour

act_duration = [[],[],[],[],[],[]]
mean_duration = [[],[],[],[],[],[]]
num_subjects = [ 0, 0, 0 ,0 ,0]

def make_hypnogram_to_time(hypnogram, event_start):
    hypnogram = hypnogram.tolist()
    for i in range(len(hypnogram)):
        hypnogram[i][0] = event_start + dt.timedelta(seconds=int(hypnogram[i][0]))
        hypnogram[i][1] = event_start + dt.timedelta(seconds=int(hypnogram[i][1]))
        hypnogram[i][2] = hypnogram[i][2]
    return hypnogram

def make_stage_event2(event_path, data_start):
    stage_label = {"N/A":7,"W":6,"S0":6,"N1":4,"S1":4,"N2":3,"S2":3,"N3":2,"S3":2,"R":5,"REM":5, "MT":1}
    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = data_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return
    #################### Event Start ###############################
    df = pd.read_excel(event_path, skiprows=row_ind-1)
    try:
        now = df[param[2]].tolist()[0]
        event_start_s = day_start +" "+ str(now.hour)+":"+str(now.minute)+":"+str(now.second)
    except:
        param[2] = "Time [hh:mm:ss]"
        now = df[param[2]].tolist()[0]
        event_start_s = day_start +" "+ str(now.hour)+":"+str(now.minute)+":"+str(now.second)

    format= "%Y-%m-%d %H:%M:%S"

    event_start = dt.datetime.strptime(event_start_s, format).replace(tzinfo=pytz.utc)
    if event_start.hour <12 : event_start += dt.timedelta(hours=12)
    ##################### Events calibration #########################
    times_s = df[param[2]].tolist()
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ str(time_s.hour)+":"+str(time_s.minute)+":"+str(time_s.second), format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    stages = df[param[0]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
    events = []
    hypnogram = np.empty((0,3), int)
    
    for i in range(len(stages)):
        event= 7
        try:
            for key in stage_label.keys():
                if key in stages[i].upper(): event = stage_label[key]
        except:
            pass
        events.append(event)

    for i in range(len(stages)):
        if i == 0 or events[i] != events[i-1]:
            ev_start = times[i]
        elif i == len(stages)-1:
            ev_end = times[i] + duration[i]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
        elif events[i] != events[i+1]:
            ev_end = times[i+1]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
    
    ##################### REM #################################
    REM_events = np.empty((0,2), int)
    
    for hyp in hypnogram:
        if hyp[2] == 5:
            REM_events = np.append(REM_events, np.array([[hyp[0],hyp[1]]]), axis = 0)
    ##################### start param #################################         
    start_index = (event_start - data_start).seconds * 200
    if event_start < data_start : start_index = -(data_start - event_start).seconds * 200
    hypnogram = make_hypnogram_to_time(hypnogram, event_start)
    return [REM_events, start_index, event_start,hypnogram,format]
    # sleep_event : [start_time, end_time]
      
def make_stage_event(event_path, data_start):
    stage_label = {"N/A":7,"W":6,"S0":6,"N1":4,"S1":4,"N2":3,"S2":3,"N3":2,"S3":2,"R":5,"REM":5, "MT":1}
    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    ind = 1
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = data_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return
    #################### Event Start ###############################
    df = pd.read_excel(event_path, skiprows=row_ind-1)
    try:
        if df[param[2]].tolist()[0].split(" ")[0] in ":": ind = 0
        event_start_s = day_start +" "+ df[param[2]].tolist()[0].split(" ")[ind].split(".")[0]
    except:
        param[2] = "Time [hh:mm:ss]"
        if df[param[2]].tolist()[0].split(" ")[0] in ":": ind = 0
        event_start_s = day_start +" "+ df[param[2]].tolist()[0].split(" ")[ind].split(".")[0]

    format= "%Y-%m-%d %H:%M:%S"

    event_start = dt.datetime.strptime(event_start_s, format).replace(tzinfo=pytz.utc)
    if event_start.hour <12 : event_start += dt.timedelta(hours=12)
    ##################### Events calibration #########################
    times_s = df[param[2]].tolist()
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ time_s.split(" ")[ind].split(".")[0], format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    stages = df[param[0]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
    events = []
    hypnogram = np.empty((0,3), int)
    
    for i in range(len(stages)):
        event= 7
        try:
            for key in stage_label.keys():
                if key in stages[i].upper(): event = stage_label[key]
        except:
            pass
        events.append(event)

    for i in range(len(stages)):
        if i == 0 or events[i] != events[i-1]:
            ev_start = times[i]
        elif i == len(stages)-1:
            ev_end = times[i] + duration[i]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
        elif events[i] != events[i+1]:
            ev_end = times[i+1]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
    
    ##################### REM #################################
    REM_events = np.empty((0,2), int)
    
    for hyp in hypnogram:
        if hyp[2] == 5:
            REM_events = np.append(REM_events, np.array([[hyp[0],hyp[1]]]), axis = 0)
    ##################### start param #################################         
    start_index = (event_start - data_start).seconds * 200
    if event_start < data_start : start_index = -(data_start - event_start).seconds * 200
    hypnogram = make_hypnogram_to_time(hypnogram, event_start)
    return [REM_events, start_index, event_start,hypnogram,format]
    # sleep_event : [start_time, end_time]
    
def make_artifact2(event_path, event_start,format, REM):
    AHI_list = ['apnea','hypopnea']
    artifact = ['artifact_chin','artifact_rarm','artifact_larm','artifact_rleg','artifact_lleg']
    artifact2 = ['artifact_chin','artifact_rfds','artifact_lfds','artifact_rta','artifact_lta']
    lists = [[],[],[],[],[]]
    AHIs = []

    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = event_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return

    df = pd.read_excel(event_path, skiprows=row_ind-1)

    ##################### Events calibration #########################
    try:
        times_s = df[param[2]].tolist()
    except:
        param[2] ="Time [hh:mm:ss]"
        times_s = df[param[2]].tolist()
    ind = 1
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ str(time_s.hour)+":"+str(time_s.minute)+":"+str(time_s.second), format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    events = df[param[3]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
   
    for i in range(5):
        for j in range(len(events)):
            if 'arousal' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if 'rera' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact2[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            
            if i != 0: continue
            if 'snore' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if events[j].lower() in AHI_list: AHIs.append([times[j],times[j]+ duration[j]])

        
        AHI = calc_AHI(AHIs, REM)
    if AHI > 15:
        for i in range(5):
            for j in range(len(events)):
                if 'apnea' in events[j]: lists[i].append([times[j],times[j]+ duration[j]])
                if 'hypopnea' in events[j]: lists[i].append([times[j],times[j]+ duration[j]])      
                
            lists[i]= np.array(lists[i])  
    return [lists, AHI]
    # sleep_event : [start_time, end_time]
        
def make_artifact(event_path, event_start,format, REM):
    artifact = ['artifact_chin','artifact_rarm','artifact_larm','artifact_rleg','artifact_lleg']
    artifact2 = ['artifact_chin','artifact_rfds','artifact_lfds','artifact_rta','artifact_lta']
    lists = [[],[],[],[],[]]
    AHIs = []

    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = event_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return

    df = pd.read_excel(event_path, skiprows=row_ind-1)

    ##################### Events calibration #########################
    try:
        times_s = df[param[2]].tolist()
    except:
        param[2] ="Time [hh:mm:ss]"
        times_s = df[param[2]].tolist()
    ind = 1
    if ":" in times_s[0].split(" ")[0]: ind = 0
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ time_s.split(" ")[ind].split(".")[0], format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    events = df[param[3]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
   


    for i in range(5):
        for j in range(len(events)):
            if 'arousal' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if 'rera' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact2[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            
            if i != 0: continue
            if 'snore' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if 'apnea' in events[j].lower() : AHIs.append([times[j],times[j]+ duration[j]])
            if 'hypopnea' in events[j].lower() : AHIs.append([times[j],times[j]+ duration[j]])

        
    AHI = calc_AHI(AHIs, REM)
    if AHI > 15:
        for i in range(5):
            for j in range(len(events)):
                if 'apnea' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
                if 'hypopnea' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])   
        
            lists[i]= np.array(lists[i])     
    return [lists, AHI]
    # sleep_event : [start_time, end_time]
    
def clean_string(s):
    return re.sub(r'[^0-9.:]', '', s)

def make_stage_event3(event_path, data_start):
    stage_label = {"N/A":7,"W":6,"S0":6,"N1":4,"S1":4,"N2":3,"S2":3,"N3":2,"S3":2,"R":5,"REM":5, "MT":1}
    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    ind = 0
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = data_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return
    #################### Event Start ###############################
    df = pd.read_excel(event_path, skiprows=row_ind-1)
    df[param[2]] = df[param[2]].apply(lambda x: x.strftime('%H:%M:%S.%f') if isinstance(x, pd.Timestamp) else x)
    df[param[2]] = df[param[2]].apply(clean_string)
    df[param[2]] = pd.to_datetime(df[param[2]], format='%H:%M:%S.%f').dt.time
    try:
        now = df[param[2]].tolist()[0]
        try:
            event_start_s = day_start +" "+ str(now.hour)+":"+str(now.minute)+":"+str(now.second)
        except:
            event_start_s = day_start +" " + now
    except:
        param[2] = "Time [hh:mm:ss]"
        try:
            event_start_s = day_start +" "+ str(now.hour)+":"+str(now.minute)+":"+str(now.second)
        except:
            event_start_s = day_start +" " + now

    format= "%Y-%m-%d %H:%M:%S"

    event_start = dt.datetime.strptime(event_start_s, format).replace(tzinfo=pytz.utc)
    if 12 >= event_start.hour >=6 : event_start += dt.timedelta(hours=12)
    ##################### Events calibration #########################
    times_s = df[param[2]].tolist()
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ str(time_s.hour)+":"+str(time_s.minute)+":"+str(time_s.second), format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    stages = df[param[0]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
    events = []
    hypnogram = np.empty((0,3), int)
    
    for i in range(len(stages)):
        event= 7
        try:
            for key in stage_label.keys():
                if key in stages[i].upper(): event = stage_label[key]
        except:
            pass
        events.append(event)

    for i in range(len(stages)):
        if i == 0 or events[i] != events[i-1]:
            ev_start = times[i]
        elif i == len(stages)-1:
            ev_end = times[i] + duration[i]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
        elif events[i] != events[i+1]:
            ev_end = times[i+1]
            hypnogram = np.append(hypnogram, np.array([[ev_start, ev_end, events[i]]]), axis = 0)
    
    ##################### REM #################################
    REM_events = np.empty((0,2), int)
    
    for hyp in hypnogram:
        if hyp[2] == 5:
            REM_events = np.append(REM_events, np.array([[hyp[0],hyp[1]]]), axis = 0)
    ##################### start param #################################         
    start_index = (event_start - data_start).seconds * 200
    if event_start < data_start : start_index = -(data_start - event_start).seconds * 200
    hypnogram = make_hypnogram_to_time(hypnogram, event_start)
    return [REM_events, start_index, event_start,hypnogram,format]
    # sleep_event : [start_time, end_time]
    
 
    # sleep_event : [start_time, end_time]

def make_artifact3(event_path, event_start,format, REM):
    AHI_list = ['apnea','hypopnea']
    artifact = ['artifact_chin','artifact_rarm','artifact_larm','artifact_rleg','artifact_lleg']
    artifact2 = ['artifact_chin','artifact_rfds','artifact_lfds','artifact_rta','artifact_lta']
    lists = [[],[],[],[],[]]
    AHIs = []

    param = ["Sleep Stage", "Position", "Time [hh:mm:ss.xxx]", "Event", "Duration[s]"]
    wb = load_workbook(filename=event_path)
    ws = wb.active # 혹은 wb['Sheet1'] 같이 특정 시트 이름을 사용
    day_start = event_start.strftime("%Y-%m-%d")
    row_ind = None
    for row in ws.iter_rows(1, ws.max_row):
        if row[0].value == 'Sleep Stage':
            row_ind = row[0].row
            break    
    if row_ind is None: return

    df = pd.read_excel(event_path, skiprows=row_ind-1)
    df[param[2]] = df[param[2]].apply(clean_string)
    df[param[2]] = pd.to_datetime(df[param[2]], format='%H:%M:%S.%f').dt.time


    ##################### Events calibration #########################
    try:
        times_s = df[param[2]].tolist()
    except:
        param[2] ="Time [hh:mm:ss]"
        times_s = df[param[2]].tolist()
    ind = 1
    times = []
    for time_s in times_s:
        time_dt = dt.datetime.strptime(day_start +" "+ str(time_s.hour)+":"+str(time_s.minute)+":"+str(time_s.second), format).replace(tzinfo=pytz.utc)
        if time_dt.hour < 9 and time_dt < event_start:
            time_dt += dt.timedelta(hours=24)
        if time_dt < event_start:
            time_dt += dt.timedelta(hours = 12)
        
        times.append((time_dt-event_start).seconds)
    ##################### Hypnogram #################################
    events = df[param[3]].tolist()
    duration = [round(dur) for dur in df[param[4]].tolist()]
   
    for i in range(5):
        for j in range(len(events)):
            if 'arousal' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if 'rera' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if artifact2[i] in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            
            if i != 0: continue
            if 'snore' in events[j].lower(): lists[i].append([times[j],times[j]+ duration[j]])
            if 'apnea' in events[j].lower() : AHIs.append([times[j],times[j]+ duration[j]])
            if 'hypopnea' in events[j].lower() : AHIs.append([times[j],times[j]+ duration[j]])

        
        AHI = calc_AHI(AHIs, REM)
    if AHI > 15:
        for i in range(5):
            for j in range(len(events)):
                if 'apnea' in events[j]: lists[i].append([times[j],times[j]+ duration[j]])
                if 'hypopnea' in events[j]: lists[i].append([times[j],times[j]+ duration[j]])      
                
            lists[i]= np.array(lists[i])  
    return [lists, AHI]
    # sleep_event : [start_time, end_time]
  
def calc_AHI( AHIs, REMs):
    total_AHI = total_REM = 0
    for rem in REMs:
        total_REM += rem[1]-rem[0]      
        for ahi in AHIs:
            if rem[0] <= ahi[0] < ahi[1]<=rem[1]:
                total_AHI += 1
            elif rem[0] <= ahi[0] < rem[1] <= ahi[1]:
                total_AHI += 1
            elif ahi[0] <=rem[0] < rem[1] <= ahi[1]:
                total_AHI += 1
            elif ahi[0]<= rem[0]<ahi[1]<rem[1]:
                total_AHI += 1
    
    AHI = round(total_AHI /total_REM *60*60,3)
    
    return AHI
    
def make_REM_period(events):
    REM = [np.empty((0,2),int) for _ in events]
    
    for i in range(len(events)) :
        REM[i] = np.append(REM[i], np.array([events[i][0].tolist()]), axis = 0)
        for j in range(1, len(events[i])):
            if REM[i][-1][1] == events[i][j][0]:
                REM[i][-1][1] = events[i][j][1]
            else :
                REM[i] = np.append(REM[i], np.array([events[i][j].tolist()]), axis = 0)

    
    return REM

def fit_to_channel(channel,REM, l):
    fit_REM = REM.copy()
    for i in range(len(channel)):
        if channel[i] == False: fit_REM[i] = [1 for _ in range(l)]
        
    return fit_REM
    
def filter_Seperate(data,start_index, channels):
    
    pickss = [0,0,0,0,0]
    ch_names = []
    for i in range(len(channels)):
        if ("chin" in channels[i].lower()) or ("mentalis" in channels[i].lower()) or ("lower.left" in channels[i].lower()) or ("subm" in channels[i].lower()):
            pickss[0] = channels[i]
            ch_names.append(channels[i])
        elif ("leg" in channels[i].lower()) or ('tibialis' in channels[i].lower()) or ('tib-' in channels[i].lower()):
            if 'r' in channels[i].lower() :
                pickss[1] = channels[i]
                ch_names.append(channels[i])
            elif 'l' in channels[i].lower():
                pickss[2] =channels[i]
                ch_names.append(channels[i])
        elif ("arm" in channels[i].lower()) or ('fds' in channels[i].lower()):
            if 'l' in channels[i].lower(): pickss[4] =channels[i];ch_names.append(channels[i])
            else: pickss[3] = channels[i];ch_names.append(channels[i])
    
    picks = [item for item in pickss if item != 0]
    datas = data.copy().pick_channels(picks, ordered = True)
    noise_freq = 60
    datas.notch_filter(noise_freq)
    data_filt = datas.copy().filter(l_freq=10, h_freq = 80)
    data_res = data_filt.resample(200)
    f_EMG_data = data_res.get_data()   #make filtered EMG_data
    RTA_EMG = []
    LTA_EMG = []
    Rarm_EMG = []
    Larm_EMG = []
    chin_EMG = []
    EMG = [chin_EMG, RTA_EMG, LTA_EMG, Rarm_EMG, Larm_EMG  ]

    channel = [False, False, False, False, False]
    ch = ["Chin","Rt_TA","Lt_TA","Rt_FDS","Lt_FDS"]

    if start_index <0:
        f_EMG_data = np.pad(f_EMG_data, ((0, 0), (-start_index, 0)), mode='constant', constant_values=0)
        start_index = 0
    for i in range(len(picks)):
        if ("chin" in picks[i].lower()) or ("mentalis" in picks[i].lower()) or ("lower.left" in picks[i].lower()):
            if (channel[0] == False):
                EMG[0] = f_EMG_data[i][start_index:]
                channel[0] = True
                
        elif ("leg" in picks[i].lower()) or ('tibialis' in picks[i].lower()):
            if ('r' in picks[i].lower() ) :
                EMG[1] = f_EMG_data[i][start_index:]
                channel[1] = True
            else :
                EMG[2] = f_EMG_data[i][start_index:]
                channel[2] = True
                
        elif ("arm" in picks[i].lower()) or ('fds' in picks[i].lower()):
            if ('l' in picks[i].lower() ):
                EMG[4] = f_EMG_data[i][start_index:]
                channel[4] = True
            else:
                EMG[3] = f_EMG_data[i][start_index:]
                channel[3] = True           


    EMG = fit_to_channel(channel, EMG, len(f_EMG_data[0][start_index:])) # channel * timepoints
    # data_res.plot(start = int(start_index/256),time_format = 'clock')
    # input()
    trg = [True for _ in range(5)]
    if channel[1] == False or channel[2] == False:
        trg[1] = False; trg[4] = False
    if channel[3] == False or channel[4] == False:
        trg[2] = False; trg[3] = False;trg[4] = False

    return [EMG, ch, trg]

def make_REM(REMs, artifacts):
    epochs = [np.empty((0,2),int) for _ in artifacts] 

    
    for i in range(len(artifacts)):
        for R in REMs:
            t_end = start = R[0]
            end = R[1]
            
            while t_end < end :
                if (end - t_end) >= 30:
                    t_end+=30
                    epochs[i] = np.append(epochs[i], np.array([[start,t_end]]),axis =0)
                    start+=30
                else : break

    return epochs

def make_REM_epochs(REMs, artifacts):
    epochs = [np.empty((0,2),int) for _ in artifacts]
    base_trgs= [[] for _ in artifacts] 
    trgs = []
    
    for i in range(len(artifacts)):
        for R in REMs:
            t_end = start = R[0]
            end = R[1]
            
            while t_end < end :
                if (end - t_end) >= 30:
                    t_end+=30
                    trg = 1
                    for arfs in artifacts[i]:
                        if (start <= arfs[0] <= t_end): trg=0
                        if (start <= arfs[1] <= t_end): trg=0
                        if (arfs[0] <= start <= arfs[1]): trg=0
                        if (arfs[0] <= t_end <= arfs[1]): trg=0
                    
                    base_trgs[i].append(trg)
                    epochs[i] = np.append(epochs[i], np.array([[start,t_end]]),axis =0)
                    start+=30
                else : break
                
    for i in range(len(artifacts)):
        ok_base = False
        oks =0
        for trg in base_trgs[i]:
            if trg == 1:
                oks+=1
            else: 
                oks = 0
            if oks == 3:
                ok_base = True
                break
        
        if ok_base == True : 
            trgs.append(0)
            continue
        epochs[i] = np.empty((0,2),int)
        base_trgs[i] = []
        trgs.append(1)

        for R in REMs:
            t_end = start = R[0]
            end = R[1]
            
            while t_end < end :
                if (end - t_end) >= 30:
                    t_end+=30
                    trg = 1
                    for arfs in artifacts[i]:
                        if (start <= arfs[0] <= t_end): trg=0
                        if (start <= arfs[1] <= t_end): trg=0
                        if (arfs[0] <= start <= arfs[1]): trg=0
                        if (arfs[0] <= t_end <= arfs[1]): trg=0
                    
                    base_trgs[i].append(0)
                    epochs[i] = np.append(epochs[i], np.array([[start,t_end]]),axis =0)
                    start+=30
                else : break   

    return [epochs, base_trgs]

def RMS(E):
    rms = np.sqrt(np.mean(np.square(E)))
    return rms

def make_rms(EMG, epochs , f_s):
    epoch_rms = []
    for i in range(len(EMG)): 
        epoch_rms.append(np.zeros(len(epochs[i])))
    epoch_rms = np.array(epoch_rms, dtype=object)
    
    i=0

    for E in EMG:  #make rms
        j=0
        
        for epoch in epochs[i]:
            epoch_start = epoch[0] * f_s
            epoch_end   = epoch[1] * f_s

            rms = RMS(E[epoch_start : epoch_end])
            if rms < 0.05e-06:
                rms = 11

            epoch_rms[i][j] = rms
            
            j+=1
        

        i+=1
    return epoch_rms

def make_baseline(epoch_rms, epochs, art_epochs):
    baselines = []
    res_baselines = []
    for i in range(len(epochs)):
        baselines.append(np.ones(len(epochs[i])))
        res_baselines.append(np.ones(len(epochs[i])))
    baselines = np.array(baselines, dtype= object)
    res_baselines = np.array(baselines, dtype= object)
    
    period= 90

    for i in range(len(epochs)):
        length =  epochs[i][0][1] - epochs[i][0][0] 
        num = period // length
        eps = []
        for j in range(len(epochs[i])):
            if art_epochs[i][j] ==1:
                eps.append(j)

        ind = 0
        for j in range(1,len(eps)):
            if epochs[i][eps[j]][0] - epochs[i][eps[j-1]][1] > 1:
                if eps[j] - ind >= num:
                    base = np.min(epoch_rms[i][ind:eps[j]])
                    for k in range(ind,eps[j]):
                        baselines[i][k] = base
                ind = eps[j]
                
        if ind == 0:
            for k in range(len(baselines[i])):
                if len(eps) == 0: baselines[i][k] = 11; continue
                baselines[i][k] = np.min(epoch_rms[i])
    l = 0
    for baseline in baselines:
        for i in range(len(baseline)):
            if baseline[i] == 1:
                L_index = len(epochs)-1
                R_index = 0
                for j in range(len(baseline)):
                    if   j<i and baseline[j] != 1 : L_index = j
                    elif j>i and baseline[j] != 1 : R_index = j; break
                diff_L = epochs[l][i][0] - epochs[l][L_index][1]
                diff_R = epochs[l][R_index][0] - epochs[l][i][1]
                
                if diff_R > diff_L:
                    res_baselines[l][i] = baseline[R_index]
                elif diff_L > diff_R:
                    res_baselines[l][i] = baseline[L_index]
                elif diff_L == diff_R:
                    res_baselines[l][i] = (baseline[R_index] + baseline[L_index])/2
            else:
                res_baselines[l][i] = baseline[i]
        l+=1
        
    
    for i in range(len(res_baselines)):
        for j in range(len(baselines)):
            if epoch_rms[i][j] == 11:
                baselines[j] = 11
        
    return res_baselines

def make_activity(baselines, artifacts, EMGs, f_s, epochs):
    """
    Detect and process 'activity' segments in EMG signals based on provided baselines, epochs, and artifacts.
    
    Parameters
    ----------
    baselines : list of lists
        Each element is a list of baseline RMS values for a given channel.
    artifacts : list of arrays
        Each element is an array of artifact time intervals (start, end).
    EMGs : list of arrays
        Each element is the EMG signal for a given channel.
    f_s : float
        Sampling frequency of the EMG signals.
    epochs : list of lists
        Each element is a list of (start, end) epoch times for that channel.
        
    Returns
    -------
    final_activities : list of np.arrays
        Each element contains detected activity segments [start_idx, end_idx, 1]
        after merging and filtering for artifacts.
    """

    # Define analysis window step size: 0.015 seconds * f_s samples
    step_size = int(0.015 * f_s)
    # Define minimal initial detection window size: 0.03 seconds * f_s samples
    initial_bouts = int(0.03 * f_s)
    count =[0,0,0,0,0]
    r_count =[0,0,0,0,0]
    
    # Initialize lists for storing detected activities at various stages
    detected_segments = [np.empty((0, 3), int) for _ in range(len(baselines))]
    merged_segments = [np.empty((0, 3), int) for _ in range(len(baselines))]
    final_activities = [np.empty((0, 3), int) for _ in range(len(baselines))]

    for ch_idx in range(len(baselines)):
        baseline_values = baselines[ch_idx]
        channel_EMG = EMGs[ch_idx]
        channel_artifacts = artifacts[ch_idx]
        channel_epochs = epochs[ch_idx]

        for epoch_idx, baseline in enumerate(baseline_values):
            # Threshold for detection
            detection_threshold = baseline * 2

            # Get epoch start and end in samples
            epoch_start = int(channel_epochs[epoch_idx][0] * f_s)
            epoch_end = int(channel_epochs[epoch_idx][1] * f_s)

            # Initialize pointers
            start_ptr = epoch_start
            end_ptr = start_ptr + initial_bouts

            # Variables to keep track of active detection phases
            up_duration = 0
            down_duration = 0
            odd_triggered = False
            activity_start = 0

            while end_ptr <= epoch_end:
                # Compute RMS in current window
                window_rms = RMS(channel_EMG[start_ptr:end_ptr])

                # If we are not yet in a prolonged active state
                if up_duration <= 0.1 * f_s:
                    if window_rms >= detection_threshold:
                        # Extend ongoing activity if close to previous segment
                        if len(detected_segments[ch_idx]) != 0 and (start_ptr - detected_segments[ch_idx][-1][1]) < 0.25 * f_s:
                            activity_start = int(start_ptr - 0.105 * f_s)
                            up_duration = int(0.105 * f_s)
                            odd_triggered = True
                        # Start a new activity segment if none is ongoing
                        if up_duration == 0:
                            activity_start = start_ptr
                            up_duration += step_size
                        else:
                            # Extend current activity segment
                            down_duration = 0
                            up_duration += step_size
                    else:
                        # Reset if RMS falls below threshold before fully establishing
                        up_duration = 0
                else:
                    # We are in a more established active period
                    if window_rms >= detection_threshold:
                        # Still active, extend activity
                        up_duration += step_size
                        down_duration = 0
                    else:
                        # Activity dropping
                        down_duration += step_size
                        # If inactivity is long enough, finalize this activity
                        if down_duration >= 0.25 * f_s:
                            # Double-check if this segment is valid
                            if not odd_triggered:
                                rms_now = RMS(channel_EMG[activity_start:end_ptr - down_duration])
                                # Check previous baseline segment if possible
                                prev_start = max(activity_start - int(0.25 * f_s), 0)
                                rms_prev = RMS(channel_EMG[prev_start:activity_start])

                                if rms_now > 2 * rms_prev:
                                    detected_segments[ch_idx] = np.append(
                                        detected_segments[ch_idx],
                                        np.array([[activity_start, end_ptr - down_duration, 1]]),
                                        axis=0
                                    )
                            else:
                                detected_segments[ch_idx] = np.append(
                                    detected_segments[ch_idx],
                                    np.array([[activity_start, end_ptr - down_duration, 1]]),
                                    axis=0
                                )

                            # Reset durations
                            up_duration = 0
                            odd_triggered = False

                # Move forward in time by step_size
                start_ptr += step_size
                end_ptr += step_size

            # If ended the epoch with ongoing activity
            if up_duration > 0.1 * f_s:
                detected_segments[ch_idx] = np.append(
                    detected_segments[ch_idx],
                    np.array([[activity_start, end_ptr - down_duration, 1]]),
                    axis=0
                )

        # Merge close segments
        if len(detected_segments[ch_idx]) > 0:
            merged_segments[ch_idx] = np.append(merged_segments[ch_idx], [detected_segments[ch_idx][0]], axis=0)
            for seg_idx in range(1, len(detected_segments[ch_idx])):
                # If next segment starts within 0.25*f_s of the last merged segment, merge them
                if detected_segments[ch_idx][seg_idx][0] - merged_segments[ch_idx][-1][1] <= 0.25 * f_s:
                    merged_segments[ch_idx][-1][1] = detected_segments[ch_idx][seg_idx][1]
                else:
                    merged_segments[ch_idx] = np.append(merged_segments[ch_idx], [detected_segments[ch_idx][seg_idx]], axis=0)

        # Filter merged segments by artifact presence and amplitude threshold
        for seg in merged_segments[ch_idx]:
            seg_start, seg_end, _ = seg
            seg_start_time = seg_start // f_s
            seg_end_time = seg_end // f_s
            is_valid = True

            # Check overlap with artifacts
            if is_valid:
                for art in channel_artifacts:
                    art_start, art_end = art
                    # Check for overlap
                    if (seg_start_time <= art_end <= seg_end_time) or \
                       (seg_start_time <= art_start <= seg_end_time) or \
                       (art_start <= seg_start_time <= art_end) or \
                       (art_start <= seg_end_time <= art_end):
                        is_valid = False
                        break
                # 파라미터
            f_s = 200  # 샘플링 레이트 가정

            if is_valid:
                # 세그먼트 신호 추출
                seg_data = channel_EMG[seg_start:seg_end]

                # FFT 계산
                N = len(seg_data)
                freqs = np.fft.rfftfreq(N, d=1/f_s)  # 양의 주파수 영역
                fft_vals = np.fft.rfft(seg_data)
                fft_mag = np.abs(fft_vals)

                k = False
                # 여기서 RWA 조건 추가
                if is_valid:
                    ranges = ((freqs >= 10) & (freqs < 55)) | ((freqs > 65) & (freqs <= 90))
                    mean_p = np.mean(fft_mag[ranges])
                    artifact_freq_range = (freqs >= 55) & (freqs <= 65)
                    if np.any(artifact_freq_range):
                        artifact_freq_max = np.max(fft_mag[artifact_freq_range])
                    else:
                        artifact_freq_max = 0
                        
                    if artifact_freq_max > 5 * mean_p:
                        is_valid = False
                
            if is_valid:
                final_activities[ch_idx] = np.append(final_activities[ch_idx], [seg], axis=0)
                r_count[ch_idx] += 1
            else:
                count[ch_idx] += 1

            count[ch_idx]/=r_count[ch_idx]+count[ch_idx]
            count[ch_idx] *= 100
    return final_activities,count
def make_event(activitys, epochs, f_s, artifacts):
    # 상수 정의
    TONIC_CUT = 5 * f_s       # tonic 판단 기준(5초)
    LONG_DURATION = 15 * f_s  # 30초 구간에서 '길게 활성' 판단 기준(15초)
    SHORT_SEG = 3 * f_s       # 3초 세그먼트 단위
    SEG_COUNT_30S = 10        # 30초를 3초로 나눴을 때 세그먼트 수 (30/3=10)

    events30 = [np.empty((0, 3), int) for _ in range(len(activitys))]
    events3 = [np.empty((0, 3), int) for _ in range(len(activitys))]

    # 30초 단위 이벤트 분류
    for ch_idx in range(len(activitys)):
        channel_activities = activitys[ch_idx]
        channel_epochs = epochs[ch_idx]
        channel_artifacts = artifacts[ch_idx]

        for epoch_idx, epoch in enumerate(channel_epochs):
            e_start = int(epoch[0] * f_s)
            e_end = int(epoch[1] * f_s)

            duration = 0
            tonic_duration = 0

            # 현재 epoch 안에서 activity 길이 계산
            for act_idx in range(len(channel_activities)):
                a_start = channel_activities[act_idx][0]
                a_end = channel_activities[act_idx][1]

                # 활동 길이가 5초 미만이면 무시
                if a_end - a_start < TONIC_CUT:
                    continue

                # epoch 내에서 활동 구간(a_start~a_end)과 epoch(e_start~e_end) 겹치는 부분 계산
                overlap_start = max(e_start, a_start)
                overlap_end = min(e_end, a_end)

                if overlap_end > overlap_start:
                    overlap_len = overlap_end - overlap_start
                    # tonic 기준 시간 이상 겹치면 tonic_duration에 추가
                    tonic_duration += overlap_len

            # 아티팩트 여부 판단
            art_trg = 0
            for art_idx in range(len(channel_artifacts)):
                # 아티팩트 시간을 샘플로 변환했다고 가정 (art[0], art[1]는 초단위라 가정)
                art = channel_artifacts[art_idx]
                art_start = int(art[0] * f_s)
                art_end = int(art[1] * f_s)
                
                overlap_start = max(e_start, a_start)
                overlap_end = min(e_end, a_end)
                # epoch 구간과 아티팩트 구간이 겹치는지 체크
                if (art_start <= e_start <= art_end) or \
                   (art_start <= e_end <= art_end) or \
                   (e_start <= art_start <= e_end):
                    art_trg = 1
                    break

            # 30초 epoch에 대한 이벤트 분류
            # duration >= 15초면 tonic
            if tonic_duration >= LONG_DURATION:
                events30[ch_idx] = np.append(events30[ch_idx], np.array([[e_start, e_end, 0]]), axis=0)
            elif art_trg == 1:
                events30[ch_idx] = np.append(events30[ch_idx], np.array([[e_start, e_end, 11]]), axis=0)
            else:
                # 이도 저도 아니면 10
                events30[ch_idx] = np.append(events30[ch_idx], np.array([[e_start, e_end, 10]]), axis=0)
                     # 아티팩트 있으면 11


    # 3초 단위 이벤트 분류
    for ch_idx in range(len(activitys)):
        channel_activities = activitys[ch_idx]
        channel_epochs = epochs[ch_idx]
        channel_artifacts = artifacts[ch_idx]

        for epoch_idx, epoch in enumerate(channel_epochs):
            e_start = int(epoch[0] * f_s)
            e_end = int(epoch[1] * f_s)

            # 3초씩 잘라서 이벤트 분류
            e_t_start = e_start
            e_t_end = e_start + SHORT_SEG

            # 30초 이벤트 결과에 따른 처리
            # 만약 30초 이벤트가 tonic이라면 3초 구간 전부 tonic으로
            if events30[ch_idx][epoch_idx][2] == 0:
                for _ in range(SEG_COUNT_30S):
                    events3[ch_idx] = np.append(events3[ch_idx], np.array([[e_t_start, e_t_end, 0]]), axis=0)
                    e_t_start += SHORT_SEG
                    e_t_end += SHORT_SEG
                # 다음 epoch으로
                continue

            # 그 외 경우 3초 단위로 세분화
            while e_t_end <= e_end:
                duration = 0
                tonic_duration = 0

                # 3초 구간 내 activity 계산
                for act_idx in range(len(channel_activities)):
                    a_start = channel_activities[act_idx][0]
                    a_end = channel_activities[act_idx][1]

                    overlap_start = max(e_t_start, a_start)
                    overlap_end = min(e_t_end, a_end)

                    if overlap_end > overlap_start:
                        overlap_len = overlap_end - overlap_start
                        
                        if a_end-a_start >= TONIC_CUT:
                            tonic_duration += overlap_len
                        else:
                            duration += overlap_len

                # 아티팩트 체크
                art_trg = 0
                for art_idx in range(len(channel_artifacts)):
                    art = channel_artifacts[art_idx]
                    art_start = int(art[0] * f_s)
                    art_end = int(art[1] * f_s)

                    if (art_start <= e_t_start <= art_end) or \
                       (art_start <= e_t_end <= art_end) or \
                       (e_t_start <= art_start <= e_t_end):
                        art_trg = 1
                        break

                # 3초 이벤트 분류
                # tonic_duration > 0 이면 intermediate
                if tonic_duration > 0:
                    events3[ch_idx] = np.append(events3[ch_idx], np.array([[e_t_start, e_t_end, 2]]), axis=0)
                # duration > 0 이면 phasic
                elif duration > 0:
                    events3[ch_idx] = np.append(events3[ch_idx], np.array([[e_t_start, e_t_end, 1]]), axis=0)
                # 아티팩트 있으면 11
                elif art_trg == 1:
                    events3[ch_idx] = np.append(events3[ch_idx], np.array([[e_t_start, e_t_end, 11]]), axis=0)
                else:
                    # 아무것도 아니면 10
                    events3[ch_idx] = np.append(events3[ch_idx], np.array([[e_t_start, e_t_end, 10]]), axis=0)

                e_t_start += SHORT_SEG
                e_t_end += SHORT_SEG

            if events30[ch_idx][epoch_idx][2] != 10: continue
            # 마지막으로 30초 이벤트 재검토
            # 마지막 10개의 3초 세그먼트에서 phasic(1) 세그먼트 카운팅
            count_phasic = 0
            count_non_idle = 0
            count_art = 0
            # 뒤에서부터 10개 세그먼트 체크 (3초 세그먼트는 10개 = 30초)
            for seg_offset in range(SEG_COUNT_30S):
                seg_event = events3[ch_idx][-seg_offset-1][2]
                if seg_event == 1:
                    count_phasic += 1
                if seg_event not in [10, 11]:  # 10=none,11=ART 제외
                    count_non_idle += 1
                if seg_event ==11:
                    count_art +=1

            # count_phasic >= 5이면 events30의 이벤트를 1(phasic)로 업데이트
            # if count_art >=5:
            #     events30[ch_idx][epoch_idx][2] = 11
            if count_phasic >= 5:
                events30[ch_idx][epoch_idx][2] = 1
            # tonic(0), phasic(1) 아닌 나머지들 중 non_idle >=5면 2(any)로 업데이트
            elif count_non_idle >= 5 and events30[ch_idx][epoch_idx][2] != 0:
                events30[ch_idx][epoch_idx][2] = 2

    return events3, events30
    
def make_activity2(EMG, f_s, event):
    bouts = int(0.03 * f_s)
    activitys = [np.empty((0,3),int) for i in range(len(EMG))]

    for j in range(len(EMG)):
        if j == 0: lm = 0
        if j == 1: lm = 1
        if j == 2: lm = 1
        if j == 3: lm = 2
        elif j == 4: lm = 2
        
        for i in range(len(event[lm])):
            if event[lm][i][2] == 0:
                ss = start = event[lm][i][0]
                t_end = start + bouts
                end   = start + 30*f_s
                up_duration = 0
                down_duration = 0
                activity_start = 0
                baseline4 = float(RMS(EMG[j][ss : end]))*2
                while t_end <= end :
                    rms = RMS(EMG[j][start : t_end])
                    if up_duration < 0.1*f_s:
                        if rms >= baseline4:
                            if up_duration == 0:
                                activity_start = start
                                up_duration += int(0.015 * f_s)
                            if activity_start != 0:
                                down_duration = 0
                                up_duration += int(0.015 * f_s)
                        else:
                            up_duration = 0
                    else:
                        if rms >= baseline4:
                            up_duration += int(0.015 * f_s)
                            down_duration = 0
                        else:
                            down_duration += int(0.015 * f_s)
                            if down_duration >= 0.25*f_s:
                                activitys[j] = np.append(activitys[j], np.array([[activity_start, t_end-down_duration, 1]]), axis=0)
                                up_duration = 0

                    start += int(0.015 * f_s)
                    t_end += int(0.015 * f_s)

                if up_duration >= 0.1*f_s:
                    activitys[j] = np.append(activitys[j], np.array([[activity_start, t_end-down_duration, 1]]), axis=0)
                i += 10

        valid_segments = np.empty((0,3), int)
        for seg in activitys[j]:
            seg_start, seg_end, _ = seg
            seg_data = EMG[j][seg_start:seg_end]
            N = len(seg_data)
            if N < 1:
                continue

            freqs = np.fft.rfftfreq(N, d=1/f_s)
            fft_vals = np.fft.rfft(seg_data)
            fft_mag = np.abs(fft_vals)

            ranges = ((freqs >= 10) & (freqs < 55)) | ((freqs > 65) & (freqs <= 100))
            if np.any(ranges):
                mean_p = np.mean(fft_mag[ranges])
            else:
                mean_p = 0

            artifact_freq_range = (freqs >= 55) & (freqs <= 65)
            if np.any(artifact_freq_range):
                artifact_freq_max = np.max(fft_mag[artifact_freq_range])
            else:
                artifact_freq_max = 0

            # 아티팩트 판단: 55~65Hz 대역 최대값이 mean_p의 5배 초과하면 아티팩트
            is_valid = True
            if mean_p > 0 and (artifact_freq_max > 5 * mean_p):
                is_valid = False

            if is_valid:
                valid_segments = np.append(valid_segments, [seg], axis=0)

        activitys[j] = valid_segments

        # 인접하거나 겹치는 세그먼트 병합 처리
        if len(activitys[j]) > 0:
            # 시작 시간 기준으로 정렬 (혹시 정렬 안되어 있을 경우)
            activitys[j] = activitys[j][activitys[j][:,0].argsort()]

            merged_segments = np.empty((0,3), int)
            merged_segments = np.append(merged_segments, [activitys[j][0]], axis=0)

            for seg_idx in range(1, len(activitys[j])):
                # 다음 세그먼트의 시작점이 이전 세그먼트의 끝점으로부터 0.25초 내이면 병합
                if (activitys[j][seg_idx][0] - merged_segments[-1][1]) <= 0.25 * f_s:
                    # 이전 병합 세그먼트의 끝점을 현재 세그먼트 끝점으로 업데이트
                    merged_segments[-1][1] = activitys[j][seg_idx][1]
                else:
                    merged_segments = np.append(merged_segments, [activitys[j][seg_idx]], axis=0)
            
            activitys[j] = merged_segments

    return activitys

def make_event2(activitys, f_s,event): 
    for i in range(len(event)): 
        for e in event[i]:
            if e[2] == 0:
                e_start = e[0]
                e_end   = e[1]
                duration = 0
                for j in range(len(activitys[i])):
                    a_start = activitys[i][j][0]
                    a_end   = activitys[i][j][1]   
                    if  a_start <= e_start < a_end <= e_end :
                        duration += a_end - e_start
                    if e_start < a_start < a_end <= e_end  :
                        duration += a_end - a_start               
                    if e_start <= a_start < e_end <= a_end :
                        duration += e_end - a_start         
                    if a_start <= e_start < e_end <= a_end :
                        duration += e_end - e_start
                    
                if 5*f_s>= duration >=20 :  e[2] = 3
                elif 5*f_s < duration :  e[2] = 4
                else : e[2] = 0
        i+=1        
    return event

def data_for_plot(events, start , f_s):
    datas = [[[0,0,0] for _ in range(len(events[i]))] for i in range(len(events))]
    
    for i in range(len(events)):
        for j in range(len(events[i])):
            datas[i][j][0] = start + dt.timedelta(seconds=round(events[i][j][0]/f_s))
            datas[i][j][1] = start + dt.timedelta(seconds=round(events[i][j][1]/f_s))
            datas[i][j][2] = events[i][j][2]
            
    return datas

def data_for_plot_ac(act, start, f_s):
    d_act = [[[0,0,0] for _ in range(len(act[i]))] for i in range(len(act))]
    
    for i in range(len(act)):
        for j in range(len(act[i])):
            d_act[i][j][0] = start + dt.timedelta(microseconds=int(act[i][j][0]/f_s*1000000))
            d_act[i][j][1] = start + dt.timedelta(microseconds=int(act[i][j][1]/f_s*1000000))
            d_act[i][j][2] = round((act[i][j][1] - act[i][j][0])/f_s,2)
            
    return d_act
    
def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def export_plot_data_xml(datas, hypnogram, RWA, channel,path):
    hx = np.array([h[0] for h in hypnogram])
    hy = np.array([h[2] for h in hypnogram])
    
    hy_ticks_labels = ["","MT","N3","N2","N1","R","W","unscored"]

    fig = plt.figure(figsize = (15,10))
    gs = gridspec.GridSpec(nrows= 2, ncols =1, height_ratios=[2,1])
    hyp = plt.subplot(gs[0])
    hyp.step(hx,hy,'royalblue',where= 'post')
    rem = np.ma.masked_where(hy != 5 , hy)
    hyp.step(hx,rem,'r',where= 'post')
    hyp.set_xlim([hx[0],hx[-1]])
    hyp.set_ylim(0,7)
    hyp.set_yticklabels(range(0,7))
    hyp.set_yticklabels(hy_ticks_labels)
    hyp.xaxis.set_major_locator(mdates.HourLocator())
    hyp.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    hyp.set_ylabel("Stage", fontsize = 10)
    hyp.set_title(channel)
    
    y_ticks_labels = ["","Phasic","Intermediate","Tonic",""]
    ax = range(len(datas))

    x = np.array([d[0] for d in datas])
    y = np.array([d[2] for d in datas])
    t = "Tonic : " + str(round(RWA[0]*100,3)) +" %"
    i = "Intermediate : " + str(round(RWA[1]*100,3)) +" %"
    p = "Phasic : " + str(round(RWA[2]*100,3)) +" %"

    
    ax = plt.subplot(gs[1])
    phasic = np.ma.masked_where(y != 1 , y)
    tonic = np.ma.masked_where(y != 3 , y)
    inter = np.ma.masked_where(y != 2 , y)
    ax.plot(x,tonic, 'bs', markersize = 4.5, label =t )
    ax.plot(x,inter, 'gs', markersize = 4.5, label = i)
    ax.plot(x,phasic, 'ys', markersize = 4.5, label = p)
    

    ax.set_xlim([hx[0],hx[-1]])
    ax.set_yticks(range(0,5), labels = y_ticks_labels)
    colors = ['w','y','g','b','w']
    for ytick, color in zip(ax.get_yticklabels(), colors):
        ytick.set_color(color)
    ax.xaxis.set_major_locator(mdates.HourLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    ax.grid(axis ='y')
    ax.set_xlabel("Time (h:m:s)", fontsize = 10)
    ax.set_ylabel("Event" , fontsize = 10)
    ax.legend(loc = 'upper left')
    fig.tight_layout()

    for d in datas:
        d[0] = d[0].strftime("%H:%M:%S.%f")
        d[1] = d[1].strftime("%H:%M:%S.%f")
        if   d[2] == 1 : d[2] = 'Phasic'
        elif d[2] == 0 : d[2] = 'Tonic'
        elif d[2] == 2 : d[2] = 'Intermediate'
        elif d[2] == 3 : d[2] = 'Tonic / Phasic'
        elif d[2] == 4 : d[2] = "Tonic / Inter"
        elif d[2] == 11: d[2] = 'Artifact'
        elif d[2] == 10: d[2] = 'No'
        

    df = pd.DataFrame(datas, columns = [ 'Start', 'End', 'Event'])
    with pd.ExcelWriter(path, mode ='a', if_sheet_exists= "overlay", engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name =channel)
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format ='png')

    wb = xl.load_workbook(path,data_only=True)
    ws = wb[channel]
    im = PILImage.open(imgdata)
    pil_img = im.resize((1200,700))
    pil_img.save("resiged.png")
    xl_img = XLImage("resiged.png")
    ws.add_image(xl_img, "F6")

    wb.save(path)
    wb.close()
    plt.close() 
    gc.collect()
    
def get_ac_RWA(act, epoch,fs):

    ac_RWA = [0 for _ in range(len(epoch))]
    for i in range(len(epoch)):
        tot_rem = 0
        tot_ac  = 0
        for rem in epoch[i] : tot_rem += (rem[1]-rem[0])
        for ac  in act[i] : tot_ac  += ( ac[1]- ac[0])
        ac_RWA[i] = tot_ac/fs/tot_rem
    return ac_RWA

def export_plot_AASM_xml(datas, hypnogram, RWA, channel,path):
    hx = np.array([h[0] for h in hypnogram])
    hy = np.array([h[2] for h in hypnogram])
    
    hy_ticks_labels = ["","MT","N3","N2","N1","R","W","unscored"]

    fig = plt.figure(figsize = (15,10))
    gs = gridspec.GridSpec(nrows= 2, ncols =1, height_ratios=[2,1])
    hyp = plt.subplot(gs[0])
    hyp.step(hx,hy,'royalblue',where= 'post')
    rem = np.ma.masked_where(hy != 5 , hy)
    hyp.step(hx,rem,'r',where= 'post')
    hyp.set_xlim([hx[0],hx[-1]])
    hyp.set_ylim(0,7)
    hyp.set_yticklabels(range(0,7))
    hyp.set_yticklabels(hy_ticks_labels)
    hyp.xaxis.set_major_locator(mdates.HourLocator())
    hyp.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    hyp.set_ylabel("Stage", fontsize = 10)
    hyp.set_title(channel)
    
    y_ticks_labels = ["","Phasic","Inter","Tonic",""]
    ax = range(len(datas))

    x = np.array([d[0] for d in datas])
    y = np.array([d[2] for d in datas])
    t = "Tonic : " + str(round(RWA[0]*100,3)) +" %"
    p = "Phasic : " + str(round(RWA[1]*100,3)) +" %"
    i = "Any : " + str(round(RWA[2]*100,3)) +" %"

    
    ax = plt.subplot(gs[1])
    phasic = np.ma.masked_where(y != 0 , y)
    tonic = np.ma.masked_where(y != 1 , y)
    any = np.ma.masked_where(y != 2 , y)
    ax.plot(x,tonic+2, 'bs', markersize = 4.5, label =t )
    ax.plot(x,phasic+1, 'ys', markersize = 4.5, label = p)
    ax.plot(x,any, 'gs', markersize = 4.5, label = i)
    

    ax.set_xlim([hx[0],hx[-1]])
    ax.set_yticks(range(0,5), labels = y_ticks_labels)
    colors = ['w','y','g','b','w']
    for ytick, color in zip(ax.get_yticklabels(), colors):
        ytick.set_color(color)
    ax.xaxis.set_major_locator(mdates.HourLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    ax.grid(axis ='y')
    ax.set_xlabel("Time (h:m:s)", fontsize = 10)
    ax.set_ylabel("Event" , fontsize = 10)
    ax.legend(loc = 'upper left')
    fig.tight_layout()

    for d in datas:
        d[0] = d[0].strftime("%H:%M:%S.%f")
        d[1] = d[1].strftime("%H:%M:%S.%f")
        if   d[2] == 0 : d[2] = 'Tonic'
        elif d[2] == 1 : d[2] = 'Phasic'
        elif d[2] == 2 : d[2] = 'Intermediate'
        elif d[2] == 10 : d[2] = 'No event'
        elif d[2] == 11: d[2] = 'Artifact'

        

    df = pd.DataFrame(datas, columns = [ 'Start', 'End', 'Event'])
    with pd.ExcelWriter(path, mode ='a', if_sheet_exists= "overlay", engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name =channel)
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format ='png')

    wb = xl.load_workbook(path,data_only=True)
    ws = wb[channel]
    im = PILImage.open(imgdata)
    pil_img = im.resize((1200,700))
    pil_img.save("resiged.png",dpi=(330, 330))
    xl_img = XLImage("resiged.png")
    ws.add_image(xl_img, "F6")

    wb.save(path)
    wb.close()
    plt.close()   
    gc.collect()
    
def get_ac_RWA(act, epoch,fs,art_duration):

    ac_RWA = [0 for _ in range(len(epoch))]
    for i in range(len(epoch)):
        
        total_art = art_duration[i]
        tot_rem = 0
        tot_ac  = 0
        for rem in epoch[i] : 
            tot_rem += (rem[1]-rem[0])
        for ac  in act[i] : tot_ac  += ( ac[1]- ac[0])
        tot_rem -= total_art
        ac_RWA[i] = tot_ac/fs/tot_rem
    return ac_RWA

def make_RWA_metric(activity1,channel,epochs, event, activity2, art_duration):
    RWA_Event = ["Tonic", "Intermediate", "Phasic","Any"]
    ev_index = {
        0 : [0,3,4], # Tonic
        1 : [2,4], # Intermediate
        2 : [1,3], # Phasic
        3 : [0,1,2,3,4] # Any
    }
    percentile = [50, 80, 95]
    RWA_mean  = [[0 for _ in channel] for _ in RWA_Event] 
    RWA_freq  = [[0 for _ in channel] for _ in RWA_Event]
    RWA_score = [[0 for _ in channel] for _ in RWA_Event]
    
    RWA_durations = [[[] for _ in channel] for _ in RWA_Event]
    RWA_percentile = [[[0 for _ in channel] for _ in percentile]for _ in RWA_Event] 

    for j in range(len(activity1)):
        RWA_t = [0 for _ in RWA_Event]
        all_time = 0
        total_art = art_duration[j]
        for ep in epochs[j]:
            all_time += ep[1] - ep[0]
        all_time -= total_art
        all_hour = all_time / 60/60
        for k in range(len(RWA_Event)): # T I P A
            for act in activity1[j]:
                trg = 0
                for ev in event[j]: # 겹치는지
                    s= ev[0]
                    e= ev[1]
                    if ev[2] not in ev_index[k]: continue
                    if act[0]>=s and act[1]<=e:trg =2
                    elif e>=act[0]>=s and act[1]>=e:trg =2
                    elif act[0] <= s and s<=act[1]<=e:trg =2
                    elif act[0] <= s and e<=act[1]:trg =2
                if trg == 2: #겹쳤다
                    if k == 2 and (act[1]- act[0])/200 >5: continue
                    if (k == 0 or k == 1) and (act[1]- act[0])/200 <=5: continue
                    RWA_durations[k][j].append(round((act[1]- act[0])/200,3))
                    RWA_mean[k][j] += (act[1]- act[0])/200 #RWA_sum
                    RWA_t[k] += 1

            for act in activity2[j]:
                trg = 0
                for ev in event[j]: # 겹치는지
                    s= ev[0]
                    e= ev[1]
                    if ev[2] not in ev_index[k]: continue
                    if act[0]>=s and act[1]<=e:trg =2
                    elif e>=act[0]>=s and act[1]>=e:trg =2
                    elif act[0] <= s and s<=act[1]<=e:trg =2
                    elif act[0] <= s and e<=act[1]:trg =2
                if trg == 2: #겹쳤다
                    if k == 2 and (act[1]- act[0])/200 >5: continue
                    if (k == 0 or k == 1) and (act[1]- act[0])/200 <=5: continue
                    RWA_durations[k][j].append(round((act[1]- act[0])/200,3))
                    RWA_mean[k][j] += (act[1]- act[0])/200 #RWA_sum
                    RWA_t[k] += 1
                           
            if RWA_t[k] != 0:
                
                RWA_score[k][j] = RWA_mean[k][j] / all_time * 100
                RWA_score[k][j] = round(RWA_score[k][j],3) 
                
                    
                RWA_mean[k][j] /= RWA_t[k]
                RWA_mean[k][j] = round(RWA_mean[k][j],3)
                
                RWA_freq[k][j] = RWA_t[k]/ all_hour
                RWA_freq[k][j] = round(RWA_freq[k][j], 3)    
                
                if all_time < 600:
                    RWA_score[k][j]  = 0     
                    RWA_mean[k][j]  = 0
                    RWA_freq[k][j]  = 0          
    
    for k in range(len(RWA_Event)):
        for p in range(len(percentile)):
            for j in range(len(activity1)):
                std_act = sorted(RWA_durations[k][j])
                n_percent = math.floor(len(std_act)*percentile[p]/100)
                if len(std_act) == 0:
                    RWA_percentile[k][p][j] = 0
                else:
                    RWA_percentile[k][p][j] = std_act[n_percent]
        
                if all_time < 600:
                    RWA_score[k][p][j]  = 0     
                    RWA_mean[k][p][j]  = 0
                    RWA_freq[k][p][j]  = 0 
    return RWA_mean , RWA_freq, RWA_score, RWA_percentile

def write_CRWA(path, hypnogram, RWA_score,RWA_mean,RWA_percentile,RWA_freq,d_act, AHI, channel):
    result_path = path+'/comb_CRWA.xlsx'
    wb = xl.Workbook()
    wb.save(result_path)
    wb.close()
    for i in range(len(d_act)):
        export_plot_act_data_xml(d_act[i], hypnogram, RWA_score[-1][i], channel[i],result_path)

            
    gc.collect()
    wb = xl.load_workbook(result_path,data_only=True)
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet("RWA summary")
    ws["A1"] = "( % )"
    ws["H1"] = "REM AHI"
    ws["H2"] = AHI
    
    for p in range(len(para)):
        ind = "A"+str(p+2)
        ws[ind] = para[p]
    for j in range(len(RWA_score[0])):
        r = 2
        ws.cell(1, j+2).value = channel[j]
        for e in range(len(RWA_Event)):
            ws.cell(r,j+2).value = RWA_score[e][j]; r+=1
            ws.cell(r,j+2).value = RWA_mean[e][j] ; r+=1
            for p in range(len(percentile)):
                ws.cell(r,j+2).value = RWA_percentile[e][p][j] 
                r+=1
            ws.cell(r,j+2).value = RWA_freq[e][j]
            r+=1
    wb.save(result_path)
    wb.close()   
    gc.collect()
    
def export_plot_act_data_xml(act, hypnogram, RWA_ac, ch,path):
    hx = np.array([h[0] for h in hypnogram])
    hx_e = np.array([h[1] for h in hypnogram])
    hy = np.array([h[2] for h in hypnogram])
    hy_ticks_labels = ["","MT","N3","N2","N1","R","W","unscored"]

    fig = plt.figure(figsize = (15,10))
    gs = gridspec.GridSpec(nrows= 2, ncols =1, height_ratios=[2,1])
    hyp = plt.subplot(gs[0])
    hyp.step(hx,hy,'royalblue',where= 'post')
    rem = np.ma.masked_where(hy != 5 , hy)
    hyp.step(hx,rem,'r',where= 'post')
    hyp.set_xlim([hx[0],hx[-1]])
    hyp.set_ylim(0,7)
    hyp.set_yticklabels(range(0,7))
    hyp.set_yticklabels(hy_ticks_labels)
    hyp.xaxis.set_major_locator(mdates.HourLocator())
    hyp.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    hyp.set_ylabel("Stage", fontsize = 10)
    hyp.set_title(ch)
    
    y_ticks_labels = ["","activity",""]
    ax = range(len(act))

    x = np.array([d[0] for d in act])
    t = "activity : " + str(round(RWA_ac,3)) +" %"


    ax = plt.subplot(gs[1])
    y = np.ones(len(x))
    ax.plot(x,y, 'gs', markersize = 1.5, label = t)
    

    ax.set_xlim([hx[0],hx[-1]])
    ax.set_yticks(range(0,3), labels = y_ticks_labels)
    colors = ['w','g','w']
    for ytick, color in zip(ax.get_yticklabels(), colors):
        ytick.set_color(color)
    ax.xaxis.set_major_locator(mdates.HourLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    ax.grid(axis ='y')
    ax.set_xlabel("Time (h:m:s)", fontsize = 10)
    ax.set_ylabel("Event" , fontsize = 10)
    ax.legend(loc = 'upper left')
    fig.tight_layout()

    for d in act:
        d[0] = d[0].strftime("%H:%M:%S.%f")
        d[1] = d[1].strftime("%H:%M:%S.%f")
        

    df = pd.DataFrame(act, columns = ['Start', 'End', 'Duration'])
    with pd.ExcelWriter(path, mode ='a', if_sheet_exists= "overlay", engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name =ch)
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format ='png')

    wb = xl.load_workbook(path,data_only=True)
    ws = wb[ch]
    im = PILImage.open(imgdata)
    pil_img = im.resize((1200,700))
    pil_img.save("resiged.png",dpi=(330, 330))
    xl_img = XLImage("resiged.png")
    ws.add_image(xl_img, "F6")

    wb.save(path)
    wb.close()  
    plt.close()
    gc.collect()
    
def export_plot_RAI_xml(act, hypnogram, RAI, ch,path):
    hx = np.array([h[0] for h in hypnogram])
    hy = np.array([h[2] for h in hypnogram])
    hy_ticks_labels = ["","MT","N3","N2","N1","R","W","unscored"]

    fig = plt.figure(figsize = (30,50))
    gs = gridspec.GridSpec(nrows= 2, ncols =1, height_ratios=[2,6])
    hyp = plt.subplot(gs[0])
    hyp.step(hx,hy,'royalblue',where= 'post')
    rem = np.ma.masked_where(hy != 5 , hy)
    hyp.step(hx,rem,'r',where= 'post')
    hyp.set_xlim([hx[0],hx[-1]])
    hyp.set_ylim(0,7)
    hyp.set_yticklabels(range(0,7))
    hyp.set_yticklabels(hy_ticks_labels)
    hyp.xaxis.set_major_locator(mdates.HourLocator())
    hyp.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    hyp.set_ylabel("Stage", fontsize = 30)
    hyp.set_title(ch)
    
    y_ticks_labels = ["","1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20",""]
    ax = range(len(act))

    for i in range(1,21):
        x = np.array([d[0] for d in act if d[2] == i])

        ax = plt.subplot(gs[1])
        y = [ i for _ in x]
        ax.plot(x,y, 'gs', markersize = 3)
    
    ax.set_xlim([hx[0],hx[-1]])
    ax.set_yticks(range(0,22), labels = y_ticks_labels)
    colors = ['w'] +['g'] *20 +['w']
    for ytick, color in zip(ax.get_yticklabels(), colors):
        ytick.set_color(color)
    ax.xaxis.set_major_locator(mdates.HourLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    ax.grid(axis ='y')
    ax.set_xlabel("Time (h:m:s)", fontsize = 30)
    ax.set_ylabel("Amplitude (uV)" , fontsize = 30)
    ax.legend(loc = 'upper left')
    fig.tight_layout()

    for d in act:
        d[0] = d[0].strftime("%H:%M:%S.%f")
        d[1] = d[1].strftime("%H:%M:%S.%f")
        

    df = pd.DataFrame(act, columns = ['Start', 'End', 'uV'])
    with pd.ExcelWriter(path, mode ='a', if_sheet_exists= "overlay", engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name =ch)
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format ='png')

    wb = xl.load_workbook(path,data_only=True)
    ws = wb[ch]
    im = PILImage.open(imgdata)
    pil_img = im.resize((1800,1800))
    pil_img.save("resiged.png",dpi=(330, 330))
    xl_img = XLImage("resiged.png")
    ws.add_image(xl_img, "F6")

    wb.save(path)
    wb.close()  
    plt.close()
    gc.collect()
     
def write_SINBAR(path, hypnogram, d_event, channel, RWA):
    result_path = path+'/comb_SINBAR.xlsx'
    wb = xl.Workbook()
    wb.save(result_path)
    wb.close()
    for i in range(len(d_event)):
        rwa = [0,0,0,0]
        for j in range(4):
            rwa[j] = RWA[i][j]
        
        export_plot_data_xml(d_event[i], hypnogram, rwa, channel[i],result_path)
    gc.collect()
    wb = xl.load_workbook(result_path,data_only=True)
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet("RWA summary")
    ws["A1"] = "( % )"
    ws["A2"] ="Tonic"
    ws["A3"] ="Intermediate"
    ws["A4"] ="Phasic"
    ws["A5"] ="Any"
    
    for i in range(4):
        for j in range(len(RWA)):
            ws.cell(1   , j+2).value = channel[j]
            ws.cell(i+2 , j+2).value = round(RWA[j][i]*100,3)
    
    wb.save(result_path)
    wb.close()
    gc.collect()
    
def write_AASM(path, hypnogram, d_event, channel, RWA):
    result_path = path+'/comb_AASM.xlsx'
    wb = xl.Workbook()
    wb.save(result_path)
    wb.close()
    for i in range(len(d_event)):
        rwa = [0,0,0]
        for j in range(3):
            rwa[j] = RWA[i][j]
        
        export_plot_AASM_xml(d_event[i], hypnogram, rwa, channel[i],result_path)
    gc.collect()
    wb = xl.load_workbook(result_path,data_only=True)
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet("RWA summary")
    ws["A1"] = "( % )"
    ws["A2"] ="Tonic"
    ws["A3"] ="Phasic"
    ws["A4"] ="Any"
    
    for i in range(3):
        for j in range(len(RWA)):
            ws.cell(1   , j+2).value = channel[j]
            ws.cell(i+2 , j+2).value = round(RWA[j][i]*100,3)
    
    wb.save(result_path)
    wb.close()
    gc.collect()

def write_RAI(path, hypnogram, d_event, channel, RAI):
    result_path = path+'/comb_RAI.xlsx'
    wb = xl.Workbook()
    wb.save(result_path)
    wb.close()
    for i in range(len(d_event)):
        
        export_plot_RAI_xml(d_event[i], hypnogram, RAI[i], channel[i],result_path)
    gc.collect()
    wb = xl.load_workbook(result_path,data_only=True)
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet("RWA summary")
    ws["A1"] = "( % )"
    ws["A2"] ="RAI"

    for j in range(len(RAI)):
        ws.cell(1   , j+2).value = channel[j]
        ws.cell(2 , j+2).value = round(RAI[j],3)
    
    wb.save(result_path)
    wb.close()
    gc.collect()

def make_RWA(event, channel, RWA_AASM): #ch epoch event
    RWA = [[0,0,0,0] for _ in channel]
    
    for i in range(len(channel)):
        art_count = np.sum(np.array(event[i]) == 11)
        for j in range(4):
            ev_ind = ev_index[j]
            for ev in event[i]:
                if ev[2] in ev_ind:
                    RWA[i][j] +=1
            RWA[i][j] /=(len(event[i]) - art_count)
            if j == 0:
                RWA[i][j] = RWA_AASM[i][j]            
    return RWA

def make_RWA_AASM(event, channel): #ch epoch event
    RWA = [[0,0,0] for _ in channel]
    
    for i in range(len(channel)):
        art_count = np.sum(np.array(event[i]) == 11)
        for j in range(3):
            ev_ind = ev_index_AASM[j]
            for ev in event[i]:
                if ev[2] in ev_ind:
                    RWA[i][j] +=1
            RWA[i][j] /=(len(event[i]) - art_count)
            
    return RWA

def make_event3(event):
    events3  = [np.empty((0,3),int) for _ in range(len(event))]
    
    for i in range(len(event)):
        for j in range(len(event[i])):
            if event[i][j][2] != 10:
                events3[i] = np.append(events3[i],np.array([[event[i][j][0],event[i][j][1],event[i][j][2]]]),axis =0 )
                
    return events3
    
def make_RAI(EMGs, f_s, epochs, artifacts):
    bouts = int(1*f_s)
    rais = [1000 for _ in epochs]
    activity_RA = []
    
    for i in range(len(epochs)):
        es = []
        RA = []
        for j in range(len(epochs[i])):
            start = epochs[i][j][0]*f_s
            end = epochs[i][j][1]*f_s
            t_end = start + bouts
            
            while t_end <= end: 
                ##############################
                ## artifact 제외
                art_t = 0
                for art in artifacts[i]:
                    if start <= art[0]*f_s <= t_end:
                        art_t= 1; break
                    elif art[0]*f_s <= start <= art[1]*f_s:
                        art_t= 1; break
                
                if art_t == 1:
                    RA.append([start, t_end, -1])       
                    t_end += bouts
                    start += bouts     
                    continue
                ###############################
                ## 1s mini-epoch 평균 값에서 baseline 값 뺌
                e = np.mean(np.abs((EMGs[i][start: t_end])))
                st = start - 30*bouts
                ed = st + bouts
                bases = []
                while ed <= t_end+30*bouts:
                    bases.append(np.mean(np.abs((EMGs[i][st: ed]))))
                    st += bouts
                    ed += bouts
                e -= np.min(np.array(bases)) # baseline
                ###############################
                ## uV에 따라 분리
                if e < 20e-06: 
                    es.append(int(e // 1e-06)+1)
                    RA.append([start, t_end, int(e//1e-06)+1])
                else :
                    es.append(20)
                    RA.append([start, t_end, 20])
                ##############################
                t_end += bouts
                start += bouts
        activity_RA.append(np.array(RA))
        es = np.array(es)
        ## 1uV 이하 값의 비율 == RAI
        art_num = np.sum(es == 2) + np.sum(es == -1)
        rais[i] = np.sum(es == 1) / (len(es) - art_num)
    
    a= 1
    return np.array(activity_RA), rais

def get_art_duration(artifacts,epochs):
    dur = [0,0,0,0,0]
    for i in range(3):
        for epoch in epochs[i]:
            start = epoch[0]
            t_end = start + 1
            R = [start, t_end]
            end = epoch[1]
            while t_end <= end:
                trg = 0
                for art in artifacts[i]:
                    if art[0] < R[0] < R[1] < art[1]: trg = 1
                    elif art[0] < R[0] < art[1] < R[1] :  trg = 1
                    elif R[0] < art[0] < R[1] < art[1] :  trg = 1
                    elif R[0] < art[0] < art[1] < R[1] :  trg = 1
                if trg == 1:
                    dur[i] += 1
                start += 1
                t_end += 1
                R = [start, t_end]
        a = 1
    return dur
        
def merge_arts(A,B):
    # 모든 이벤트를 하나의 리스트로 합침
    A_start = []; A_end = []
    for a in A:
        A_start.append(a[0])
        A_end.append(a[1])

    B_start = []; B_end = []
    for a in B:
        B_start.append(a[0])
        B_end.append(a[1])
        
    events = []
    for start, end in zip(A_start, A_end):
        events.append((start, 'start'))
        events.append((end, 'end'))
    for start, end in zip(B_start, B_end):
        events.append((start, 'start'))
        events.append((end, 'end'))
    
    # 이벤트를 시간순으로 정렬
    events.sort()
    
    merged = []
    
    active_events = 0
    current_start = None
    
    for time, type in events:
        if type == 'start':
            if active_events == 0:
                current_start = time
            active_events += 1
        elif type == 'end':
            active_events -= 1
            if active_events == 0:
                merged.append([current_start, time])
    
    return np.array(merged)


def merge_events(A, B, tolerance=50):
    """
    A와 B에 포함된 이벤트 (start, end) 구간을 병합하는 함수.
    이벤트가 겹치거나, 겹치지 않더라도 시작점과 끝점 사이 간격이 'tolerance' 이하인 경우 하나로 병합한다.
    
    Parameters
    ----------
    A, B : list of lists or arrays
        각 원소는 [start, end] 형태의 이벤트 구간
    tolerance : float or int
        두 이벤트 사이의 최대 허용 간격. 이 간격 이하이면 하나의 이벤트로 병합.

    Returns
    -------
    np.array
        [start, end, 1] 형태의 병합된 이벤트 배열
    """
    
    A_start = []; A_end = []
    for a in A:
        A_start.append(a[0])
        A_end.append(a[1])

    B_start = []; B_end = []
    for b in B:
        B_start.append(b[0])
        B_end.append(b[1])
        
    events = []
    for start, end in zip(A_start, A_end):
        events.append((start, 'start'))
        events.append((end, 'end'))
    for start, end in zip(B_start, B_end):
        events.append((start, 'start'))
        events.append((end, 'end'))
    
    # 이벤트를 시간순으로 정렬
    events.sort()
    
    merged = []
    active_events = 0
    current_start = None
    
    # 겹치는 이벤트 처리(기존 로직)
    for time, etype in events:
        if etype == 'start':
            if active_events == 0:
                current_start = time
            active_events += 1
        elif etype == 'end':
            active_events -= 1
            if active_events == 0:
                merged.append([current_start, time, 1])
    
    # 이제 merged에는 겹치는 이벤트가 하나로 합쳐진 상태
    # 여기서 tolerance 이내로 근접한 이벤트를 추가로 병합
    merged = np.array(merged)
    if len(merged) == 0:
        return merged
    
    final_merged = [merged[0]]
    for i in range(1, len(merged)):
        # 이전 이벤트의 끝과 현재 이벤트의 시작 사이의 간격 확인
        if merged[i][0] - final_merged[-1][1] <= tolerance:
            # 두 이벤트를 병합: 기존 이벤트의 끝 시간을 현재 이벤트의 끝 시간으로 확장
            final_merged[-1][1] = merged[i][1]
        else:
            final_merged.append(merged[i])
    
    return np.array(final_merged)

def combine_act(acts,triggers):
    res = [[] for _ in [0,1,2,3,4]]
    for i in [0,1,2]:
        if i == 0: res[i] = np.array(acts[0]); continue
        elif i ==1: 
            a = acts[1]; b = acts[2]
        elif i ==2:
            a = acts[3]; b = acts[4]
        
        res[i] = merge_events(a,b)
        
    for i in [3,4]:
        if i == 3: a = res[0]; b = res[2]
        elif i ==4: 
            a = res[3]; b = res[1]
        
        res[i] = merge_events(a,b)     
    
    for i in range(5):
        if triggers[i] == False:
            res[i] = np.array([])   

    ress = []        
    for i in range(5):
        temp =[]
        for r in res[i]:
            trg = 1
            for t in temp:
                if r[0] == t[0] and r[1] == t[1]:
                    trg = 0
            if trg == 0: continue
            temp.append(r)
        ress.append(np.array(temp))
    return res

def combine_art(acts,triggers):
    res = [[] for _ in [0,1,2,3,4]]
    for i in [0,1,2]:
        if i == 0: res[i] = np.array(acts[0]); continue
        elif i ==1: 
            a = acts[1]; b = acts[2]
        elif i ==2:
            a = acts[3]; b = acts[4]
        
        res[i] = merge_arts(a,b)
        
    for i in [3,4]:
        if i == 3: a = res[0]; b = res[2]
        elif i ==4: 
            a = res[3]; b = res[1]
        
        res[i] = merge_arts(a,b)     
    
    for i in range(5):
        if triggers[i] == False:
            res[i] = np.array([])   
    return res

def combine_rai(acts, rais):
    res = [[] for _ in [0,1,2,3,4]]
    rai = [0 for _ in [0,1,2,3,4]]
    for i in [0,1,2,3,4]:
        if i == 0: res[i] = acts[0]; rai[i] = rais[i]; continue
        elif i ==1: 
            a = acts[1]; b = acts[2]
        elif i ==2:
            a = acts[3]; b = acts[4]
        
        res[i], rai[i] = merge_rai(a,b)

    for i in [3,4]:
        if i == 3: a = res[0]; b = res[2]
        elif i ==4: 
            a = res[3]; b = res[1]
        
        res[i], rai[i] = merge_rai(a,b)      
    return res, np.array(rai)

def merge_rai(a,b):
    res = []
    for i in range(len(a)):
        if -1 in [a[i][2],b[i][2]]: res.append([a[i][0],a[i][1],-1])
        elif a[i][2] >= b[i][2]: res.append(a[i])
        else: res.append(b[i])
    res= np.array(res)
    art_num = np.sum(res == 2) + np.sum(res == -1)
    rai = np.sum(res == 1) / (len(res) - art_num)
    return res, rai

def comb_event(events,triggers): #ch ep [s,e,t]
    res = [[] for _ in [0,1,2,3,4]]
    for i in [0,1,2]:
        if triggers[i] == False:
            res[i] = np.array([]); continue
        if i == 0: res[i] = np.array(events[0]); continue
        elif i ==1: 
            a = events[1][:]; b = events[2][:]
        elif i ==2:
            a =events[3][:]; b = events[4][:]
        
        res[i] = merge_ev(a,b)
        
    for i in [3,4]:
        if triggers[i] == False:
            res[i] = np.array([]); continue
        if i == 3: a = res[0][:]; b = res[2][:]
        elif i ==4: 
            a = res[3][:]; b = res[1][:]
        
        res[i] = merge_ev(a,b)     
  
    return res    

def merge_ev(A, B):
    res = []
    A = np.array(A)
    B = np.array(B)

    # A와 B 모두 [s,e,t] 형태라고 가정
    for i in range(len(A)):
        matched_index = -1
        for j in range(len(B)):
            if A[i][0] == B[j][0]:
                matched_index = j
                break
        if matched_index == -1:
            # B와 시작 시간이 일치하는 이벤트를 찾지 못한 경우 다음 i로 진행
            continue

        # 깊은 복사로 A[i]의 수정이 원본에 영향 주지 않도록 함
        tt = A[i].copy()

        valA = A[i][2]
        valB = B[matched_index][2]

        # 각각의 상태값이 어떤 것인지 판단하기 쉽게 집합으로 처리
        vals = {valA, valB}
        if 11 in vals:
            tt[2] = 11
        # 조건 로직 유지: 어떤 상태값들이 있는지에 따라 tt[2] 결정
        elif (1 in vals or 2 in vals or 10 in vals) and 0 in vals:
            tt[2] = 0
        elif 10 in vals:
            # 여기서 A[i][2]나 B[matched_index][2] 중 5보다 작은 값이 있으면 그것으로 설정
            # 원본 로직 상 valA나 valB 둘 중 하나가 5 미만이면 tt[2]를 해당 값으로 함
            if valA < 5:
                tt[2] = valA
            if valB < 5:
                tt[2] = valB
        else:
            # 위 조건에 해당하지 않으면 단순히 최대값을 취함
            tt[2] = max(valA, valB)
        
        res.append(tt)

    return np.array(res)

                
